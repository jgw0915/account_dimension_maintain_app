from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,Border,Side

class TreeNode:
    def __init__(self, data,level):
        self.data = data
        self.children = []
        self.children_map = []
        self.level = level

    def add_child(self, child_node):
        if child_node.data not in self.children_map:
            self.children.append(child_node)
            self.children_map.append(child_node.data)

class Tree:
    def __init__(self, root_data):
        self.root = TreeNode(root_data,0)

    def traverse_depth_first(self, start_node=None):
        if start_node is None:
            start_node = self.root
        yield (start_node)
        for child in start_node.children:
            yield from self.traverse_depth_first(child)
    
    def find_node(self, data, start_node=None):
        if start_node is None:
            start_node = self.root
        if start_node.data == data:
            return start_node
        for child in start_node.children:
            node = self.find_node(data, child)
            if node:
                return node
        return None

wb = load_workbook(r'C:\Users\IEC1-PICST-09\Python_Project\Small Apllication\data\費用科目列表_v3.xlsx',data_only=True)
CN_COA = ('TWCO', 'PDCO', 'CQCO', 'ITC')
Foriegn_COA =('MXCO', 'USCO', 'THCO', 'CZCO')
All_COA = CN_COA+Foriegn_COA
# print(wb.sheetnames)
COA_Tree = Tree('COA')
for sheet in wb:
    if sheet.title in All_COA:
        for row in range(2,len(sheet['A'])+1):
            # COA層
            Current_COA = str(sheet[row][5].value)[:2]+' COA'
            COA_Tree.root.add_child(TreeNode(Current_COA,level=1))
            # Fixed First one Number 層
            Next_node = COA_Tree.find_node(data=Current_COA)
            if Next_node is not None:
                if sheet.title in CN_COA:
                    COA_Prefix_1 = Current_COA[:2]+'_'+str(sheet[row][1].value)[:1]+'XXX'
                else:
                    COA_Prefix_1 = Current_COA[:2]+'_'+str(sheet[row][1].value)[:1]+ 'XXXX'       
                Next_node.add_child(TreeNode(COA_Prefix_1,level=2))
            # Fixed First two Number 層
            Next_node = COA_Tree.find_node(COA_Prefix_1,Next_node)
            if Next_node is not None:
                if sheet.title in CN_COA:
                    COA_Prefix_2 = Current_COA[:2]+'_'+str(sheet[row][1].value)[:2]+'XX'
                else:
                    COA_Prefix_2 = Current_COA[:2]+'_'+str(sheet[row][1].value)[:2]+ 'XXX'       
                Next_node.add_child(TreeNode(COA_Prefix_2,level=3))
            Next_node = COA_Tree.find_node(COA_Prefix_2,Next_node)
            if Next_node is not None:
                member_name = str(sheet[row][5].value)
                if sheet.title in CN_COA:
                    CN_Alias = member_name+'_'+str(sheet[row][2].value)
                    EN_Alias = member_name+'_'+str(sheet[row][4].value)
                else:
                    CN_Alias = member_name+'_'+str(sheet[row][2].value)
                    EN_Alias = member_name+'_'+str(sheet[row][4].value)
                if 'MFT' in member_name:   
                    member_name = member_name.replace('MFT','FT')
                    CN_Alias = CN_Alias.replace('MFT','FT')
                    EN_Alias = EN_Alias.replace('MFT','FT')
                if 'NFT' in member_name:   
                    member_name = member_name.replace('NFT','FT')
                    CN_Alias = CN_Alias.replace('NFT','FT')
                    EN_Alias = EN_Alias.replace('NFT','FT')
                data = {'member_name':member_name,'CN_Alias':CN_Alias,'EN_Alias':EN_Alias}
                Next_node.add_child(TreeNode(data,level=4))
COL_Map = ['A','B','C','D','E','F','G','H']
ws = wb.create_sheet('Organized_COA_Tree')
row_count = 1
for node in COA_Tree.traverse_depth_first():
    if type(node.data) is not type({}):
        ws['{}'.format(COL_Map[node.level]+str(row_count))] = node.data
    else:
        ws['{}'.format(COL_Map[node.level]+str(row_count))] = node.data['member_name']
        ws['{}'.format(COL_Map[node.level+1]+str(row_count))] = node.data['CN_Alias']
        ws['{}'.format(COL_Map[node.level+2]+str(row_count))] = node.data['EN_Alias']
    row_count += 1
font = Font(name='Microsoft JhengHei', size=12, bold=False, italic=False)
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
for row in ws.iter_rows():
    for cell in row:
        cell.font = font
        cell.border = border
wb.save(r'C:\Users\IEC1-PICST-09\Python_Project\Small Apllication\export file\費用科目列表_v3__COA_Tree.xlsx')