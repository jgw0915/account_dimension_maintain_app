from openpyxl import Workbook
from openpyxl import load_workbook
from collections import deque
from openpyxl.styles import Font,Border,Side


class TreeNode:
    def __init__(self, data,level):
        self.data = data
        self.children = []
        self.children_map = []
        self.level = level

    def add_child(self, child_node):
        if child_node.data not in self.children_map:
            self.children.append(child_node)
            self.children_map.append(child_node.data)

class Tree:
    def __init__(self, root_data):
        self.root = TreeNode(root_data,0)

    def traverse_breadth_first(self, start_node=None):
        queue = deque()
        queue.append(self.root)

        while queue:
            node = queue.popleft()
            yield node

            for child in node.children:
                queue.append(child)
        # if start_node is None:
        #     start_node = self.root
        # print(start_node.data)
        # for child in start_node.children:
        #     self.traverse_depth_first(child)
    
    def find_node(self, data, start_node=None):
        if start_node is None:
            start_node = self.root
        if start_node.data == data:
            return start_node
        for child in start_node.children:
            node = self.find_node(data, child)
            if node:
                return node
        return None

wb = load_workbook(r'C:\Users\IEC1-PICST-09\Python_Project\Small Apllication\data\費用科目列表_v3.xlsx',data_only=True)
CN_COA = ('TWCO', 'PDCO', 'CQCO', 'ITC')
Foriegn_COA =('MXCO', 'USCO', 'THCO', 'CZCO')
duplicate_small_cat = ('耗材/消耗品','RD Material/Sample','Tooling','水電費','旅費','廠辦租金','Excess')
All_COA = CN_COA+Foriegn_COA
# print(wb.sheetnames)
Catgory_Tree = Tree({'member_name':'AC91000','CN_Alias':'SPDG by Category','EN_Alias':''})
for sheet in wb:
    if sheet.title in All_COA:
        for row in range(2,len(sheet['A'])+1):
            # Big Catgory 層
            big_Catgory_data = {'member_name':sheet[row][15].value,'CN_Alias':sheet[row][9].value,'EN_Alias':sheet[row][11].value}
            Catgory_Tree.root.add_child(TreeNode(big_Catgory_data,level=1))
            # Middle Catgory 層
            Next_node = Catgory_Tree.find_node(data=big_Catgory_data)
            if Next_node is not None:
                middle_Catgory_data = {'member_name':sheet[row][14].value,'CN_Alias':sheet[row][8].value,'EN_Alias':sheet[row][10].value}   
                Next_node.add_child(TreeNode(middle_Catgory_data,level=2))
            # Small Catgory 層
            Next_node = Catgory_Tree.find_node(middle_Catgory_data,Next_node)
            if Next_node is not None:
                if sheet[row][3].value in duplicate_small_cat:
                    small_Catgory_data =  {'member_name':sheet[row][13].value,'CN_Alias':sheet[row][3].value+'.','EN_Alias':sheet[row][12].value+'.'}    
                else:
                    small_Catgory_data =  {'member_name':sheet[row][13].value,'CN_Alias':sheet[row][3].value,'EN_Alias':sheet[row][12].value}
                Next_node.add_child(TreeNode(small_Catgory_data,level=3))
            Next_node = Catgory_Tree.find_node(small_Catgory_data,Next_node)
            # Accounting 層
            if Next_node is not None:
                member_name = str(sheet[row][5].value)
                if sheet.title in CN_COA:
                    CN_Alias = member_name+'_'+str(sheet[row][2].value)
                    EN_Alias = member_name+'_'+str(sheet[row][4].value)
                else:
                    CN_Alias = member_name+'_'+str(sheet[row][2].value)
                    EN_Alias = member_name+'_'+str(sheet[row][4].value)
                if 'MFT' in member_name:   
                    member_name = member_name.replace('MFT','FT')
                    CN_Alias = CN_Alias.replace('MFT','FT')
                    EN_Alias = EN_Alias.replace('MFT','FT')
                if 'NFT' in member_name:   
                    member_name = member_name.replace('NFT','FT')
                    CN_Alias = CN_Alias.replace('NFT','FT')
                    EN_Alias = EN_Alias.replace('NFT','FT')
                data = {'member_name':member_name,'CN_Alias':CN_Alias,'EN_Alias':EN_Alias}
                Next_node.add_child(TreeNode(data,level=4))
COL_Map = ['A','B','C','D','E','F','G','H']
ws = wb.create_sheet('Organized_SPDG_BY_CAT_IMP')
row_count = 1
for node in Catgory_Tree.traverse_breadth_first():
    if node.level != 4:
        # if node.level != 3:
        #     for child in node.children:
        #         ws['{}'.format(COL_Map[0]+str(row_count))] = child.data
        #         ws['{}'.format(COL_Map[1]+str(row_count))] = node.data
        #         row_count += 1
        # else:
        print
        for child in node.children:
            ws['{}'.format(COL_Map[0]+str(row_count))] = child.data['member_name']
            ws['{}'.format(COL_Map[1]+str(row_count))] = node.data['member_name']
            ws['{}'.format(COL_Map[2]+str(row_count))] = child.data['CN_Alias']
            ws['{}'.format(COL_Map[3]+str(row_count))] = child.data['EN_Alias'];
            row_count += 1
font = Font(name='Microsoft JhengHei', size=12, bold=False, italic=False)
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
for row in ws.iter_rows():
    for cell in row:
        cell.font = font
        cell.border = border
wb.save(r'C:\Users\IEC1-PICST-09\Python_Project\Small Apllication\export file\費用科目列表_v3_SPDG_By_Cat_IMP_BFS.xlsx')